"""Batch processing for stratigraphic thickness models."""

from __future__ import annotations

import json
import random
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from statistics import mean, pstdev
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from source.batch_schema import (
    BATCH_HEADERS,
    DEFAULT_SIGMAS,
    DEFAULT_VALUES,
    FIELD_BOUNDS,
    INPUT_SIGMA_PAIRS,
    MC_SAMPLE_COUNT_DESKTOP,
    MODEL_ID_BY_T,
    T1_FIELD_MAP,
    T1_REQUIRED,
    T1_UNUSED,
    TWO_BED_FIELD_MAP,
    TWO_BED_REQUIRED,
    WRAP_AZIMUTH_FIELDS,
    example_rows,
)
from source.models import (
    AverageThicknessInputs,
    AverageVectorInputs,
    ConcentricFoldInputs,
    EqualAngleInputs,
    MixedAverageInputs,
    OneDipInputs,
    PlungingConcentricFoldInputs,
    TopNormalInputs,
    compute_average_thickness,
    compute_average_vector,
    compute_concentric_fold,
    compute_equal_angle,
    compute_mixed_average,
    compute_one_dip,
    compute_plunging_concentric_fold,
    compute_top_normal,
)

MODEL_COMPUTE: dict[str, tuple[type, Callable[..., Any]]] = {
    "t1": (OneDipInputs, compute_one_dip),
    "t2": (AverageVectorInputs, compute_average_vector),
    "t3": (AverageThicknessInputs, compute_average_thickness),
    "t4": (MixedAverageInputs, compute_mixed_average),
    "t5": (ConcentricFoldInputs, compute_concentric_fold),
    "t6": (PlungingConcentricFoldInputs, compute_plunging_concentric_fold),
    "t7": (TopNormalInputs, compute_top_normal),
    "t8": (EqualAngleInputs, compute_equal_angle),
}

MC_OUTPUT_COLUMNS = [
    "MC_N",
    "MC_mean",
    "MC_std",
    "MC_P10",
    "MC_P25",
    "MC_P50",
    "MC_P75",
    "MC_P90",
]

RESULT_COLUMNS = [
    "T_result",
    "status",
    *MC_OUTPUT_COLUMNS,
    "error",
    "intermediate_values",
]


@dataclass
class BatchRowResult:
    row_index: int
    well_id: str
    t_number: int
    model_id: str
    inputs: dict[str, Any]
    t_result: float | None = None
    status: str = "OK"
    error: str = ""
    intermediate_values: str = ""
    warnings: tuple[str, ...] = field(default_factory=tuple)
    mc_stats: dict[str, float] | None = None
    mc_thicknesses: list[float] | None = field(default=None, repr=False)


def _normalize_header(value: Any) -> str:
    return str(value or "").strip()


def _parse_t_number(raw: Any) -> int:
    if raw is None or str(raw).strip() == "":
        raise ValueError("Missing T model number.")
    text = str(raw).strip().upper()
    match = re.fullmatch(r"T?([1-8])", text)
    if not match:
        raise ValueError(f"Invalid T value {raw!r}; use 1–8 or T1–T8.")
    return int(match.group(1))


def _cell_empty(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _parse_float(value: Any, *, column: str) -> float:
    if _cell_empty(value):
        raise ValueError(f"Missing numeric value for {column}.")
    return float(value)


def _parse_optional_float(value: Any, *, default: float = 0.0) -> float:
    if _cell_empty(value):
        return default
    return float(value)


def _field_map_for_model(model_id: str) -> dict[str, str]:
    return T1_FIELD_MAP if model_id == "t1" else TWO_BED_FIELD_MAP


def _required_value_columns(model_id: str) -> list[str]:
    return T1_REQUIRED if model_id == "t1" else TWO_BED_REQUIRED


def _validate_unused_t1_columns(row: dict[str, Any]) -> str | None:
    for col in T1_UNUSED:
        if not _cell_empty(row.get(col)):
            return f"T1 row must leave {col} blank."
    return None


def parse_batch_workbook(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [_normalize_header(h) for h in header_row]
    if not headers or all(h == "" for h in headers):
        raise ValueError("Batch workbook is missing a header row.")
    header_index = {h: i for i, h in enumerate(headers) if h}
    missing = [h for h in BATCH_HEADERS if h not in header_index]
    if missing:
        raise ValueError(f"Batch workbook missing columns: {', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if excel_row is None or all(_cell_empty(v) for v in excel_row):
            continue
        parsed: dict[str, Any] = {}
        for header in BATCH_HEADERS:
            parsed[header] = excel_row[header_index[header]]
        rows.append(parsed)
    if not rows:
        raise ValueError("Batch workbook has no data rows.")
    return rows


def _build_model_inputs(
    row: dict[str, Any], model_id: str
) -> tuple[dict[str, float], dict[str, float]]:
    field_map = _field_map_for_model(model_id)
    values: dict[str, float] = {}
    sigmas: dict[str, float] = {}
    for value_col, sigma_col in INPUT_SIGMA_PAIRS:
        field_name = field_map.get(value_col)
        if field_name is None:
            continue
        if value_col in _required_value_columns(model_id):
            values[field_name] = _parse_float(row.get(value_col), column=value_col)
        elif not _cell_empty(row.get(value_col)):
            values[field_name] = _parse_float(row.get(value_col), column=value_col)
        sigmas[field_name] = max(0.0, _parse_optional_float(row.get(sigma_col)))
    return values, sigmas


def validate_batch_row(row: dict[str, Any]) -> tuple[int, str, dict[str, float], dict[str, float]]:
    well_id = str(row.get("well_id") or "").strip()
    if not well_id:
        raise ValueError("Missing well_id.")
    t_number = _parse_t_number(row.get("T"))
    model_id = MODEL_ID_BY_T[t_number]
    if model_id == "t1":
        unused_err = _validate_unused_t1_columns(row)
        if unused_err:
            raise ValueError(unused_err)
    values, sigmas = _build_model_inputs(row, model_id)
    missing = [
        col
        for col in _required_value_columns(model_id)
        if _cell_empty(row.get(col))
    ]
    if missing:
        raise ValueError(f"Missing required columns for T{t_number}: {', '.join(missing)}")
    return t_number, model_id, values, sigmas


def _percentile(values: list[float], p: float) -> float:
    if not values:
        raise ValueError("Cannot compute percentile on empty values.")
    if len(values) == 1:
        return values[0]
    sorted_vals = sorted(values)
    index = (len(sorted_vals) - 1) * p
    low = int(index)
    high = min(low + 1, len(sorted_vals) - 1)
    frac = index - low
    return sorted_vals[low] * (1.0 - frac) + sorted_vals[high] * frac


def _sample(
    mu: float,
    sigma: float,
    field_name: str,
    *,
    rng: random.Random,
) -> float:
    if sigma <= 1e-12:
        return mu
    lo, hi = FIELD_BOUNDS[field_name]
    raw = rng.gauss(mu, sigma)
    if field_name in WRAP_AZIMUTH_FIELDS:
        width = hi - lo
        return ((raw - lo) % width) + lo
    return min(max(raw, lo), hi)


def _run_monte_carlo(
    model_id: str,
    values: dict[str, float],
    sigmas: dict[str, float],
    *,
    sample_count: int,
    seed: int | None = None,
) -> tuple[dict[str, float], list[float]] | None:
    if not any(sigmas.get(k, 0.0) > 1e-12 for k in values):
        return None
    rng = random.Random(seed)
    input_cls, compute_fn = MODEL_COMPUTE[model_id]
    thicknesses: list[float] = []
    keys = list(values.keys())
    for _ in range(sample_count):
        cur = {
            k: _sample(values[k], sigmas.get(k, 0.0), k, rng=rng)
            for k in keys
        }
        result = compute_fn(input_cls(**cur))
        thicknesses.append(float(result.true_stratigraphic_thickness))
    stats = {
        "MC_N": float(sample_count),
        "MC_mean": mean(thicknesses),
        "MC_std": pstdev(thicknesses),
        "MC_P10": _percentile(thicknesses, 0.10),
        "MC_P25": _percentile(thicknesses, 0.25),
        "MC_P50": _percentile(thicknesses, 0.50),
        "MC_P75": _percentile(thicknesses, 0.75),
        "MC_P90": _percentile(thicknesses, 0.90),
    }
    return stats, thicknesses


def _vec3_lines(prefix: str, vector: tuple[float, float, float]) -> list[str]:
    return [
        f"{prefix}_x={vector[0]:.6f}",
        f"{prefix}_y={vector[1]:.6f}",
        f"{prefix}_z={vector[2]:.6f}",
    ]


def format_intermediate_values(model_id: str, result: Any) -> str:
    lines: list[str] = []
    data = asdict(result)
    warnings = data.pop("geometry_warnings", ()) or ()
    t_key = {
        "t1": "T1",
        "t2": "T2",
        "t3": "T3",
        "t4": "T4",
        "t5": "T5",
        "t6": "T6",
        "t7": "T7",
        "t8": "T8",
    }[model_id]
    lines.append(f"{t_key}={result.true_stratigraphic_thickness:.6f}")

    skip = {"true_stratigraphic_thickness", "geometry_warnings"}
    for key, value in data.items():
        if key in skip:
            continue
        if key.endswith("_vector") and isinstance(value, (list, tuple)) and len(value) == 3:
            prefix = key.replace("_vector", "")
            prefix = {
                "ud1": "Ud1",
                "ud2": "Ud2",
                "ud2_prime": "Ud2_prime",
                "uav": "Uav",
                "ub": "Ub",
                "ndc": "Ndc",
                "ndp": "Ndp",
                "c": "Uc",
                "ub_prime": "Ub_prime",
            }.get(prefix, prefix)
            lines.extend(_vec3_lines(prefix, tuple(float(v) for v in value)))
        elif isinstance(value, bool):
            lines.append(f"{key}={int(value)}")
        elif isinstance(value, (int, float)):
            lines.append(f"{key}={float(value):.6f}")

    if warnings:
        lines.append("geometry_warnings=" + "; ".join(str(w) for w in warnings))
    # Single-line cell content so Excel does not wrap and stay illegible.
    return "; ".join(lines)


def process_batch_row(
    row_index: int,
    row: dict[str, Any],
    *,
    sample_count: int = MC_SAMPLE_COUNT_DESKTOP,
    mc_seed: int | None = None,
) -> BatchRowResult:
    inputs = {h: row.get(h, "") for h in BATCH_HEADERS}
    well_id = str(row.get("well_id") or "").strip()
    try:
        t_number, model_id, values, sigmas = validate_batch_row(row)
        input_cls, compute_fn = MODEL_COMPUTE[model_id]
        result = compute_fn(input_cls(**values))
        mc_bundle = _run_monte_carlo(
            model_id,
            values,
            sigmas,
            sample_count=sample_count,
            seed=mc_seed,
        )
        mc_stats = None
        mc_thicknesses = None
        if mc_bundle is not None:
            mc_stats, mc_thicknesses = mc_bundle
        geom_warn = getattr(result, "geometry_warnings", ()) or ()
        return BatchRowResult(
            row_index=row_index,
            well_id=well_id,
            t_number=t_number,
            model_id=model_id,
            inputs=inputs,
            t_result=float(result.true_stratigraphic_thickness),
            status="OK",
            intermediate_values=format_intermediate_values(model_id, result),
            warnings=tuple(str(w) for w in geom_warn),
            mc_stats=mc_stats,
            mc_thicknesses=mc_thicknesses,
        )
    except Exception as exc:
        t_number = 0
        model_id = ""
        try:
            t_number = _parse_t_number(row.get("T"))
            model_id = MODEL_ID_BY_T[t_number]
        except Exception:
            pass
        return BatchRowResult(
            row_index=row_index,
            well_id=well_id,
            t_number=t_number,
            model_id=model_id,
            inputs=inputs,
            status="ERROR",
            error=str(exc),
        )


def process_batch_rows(
    rows: list[dict[str, Any]],
    *,
    sample_count: int = MC_SAMPLE_COUNT_DESKTOP,
    progress_cb: Callable[[int, int, str], None] | None = None,
) -> list[BatchRowResult]:
    results: list[BatchRowResult] = []
    total = len(rows)
    for idx, row in enumerate(rows, start=1):
        if progress_cb is not None:
            well = str(row.get("well_id") or f"row {idx}")
            progress_cb(idx, total, f"Processing {well} ({idx}/{total})...")
        results.append(
            process_batch_row(
                idx,
                row,
                sample_count=sample_count,
                mc_seed=42 + idx,
            )
        )
    return results


def _write_header_row(ws, headers: list[str]) -> None:
    bold = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold


def write_batch_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "BatchInput"
    _write_header_row(ws, BATCH_HEADERS)
    for col_idx, header in enumerate(BATCH_HEADERS, start=1):
        if header in DEFAULT_VALUES:
            ws.cell(row=2, column=col_idx, value=DEFAULT_VALUES[header])
        elif header in DEFAULT_SIGMAS:
            ws.cell(row=2, column=col_idx, value=DEFAULT_SIGMAS[header])
        elif header == "T":
            ws.cell(row=2, column=col_idx, value=1)
        elif header == "well_id":
            ws.cell(row=2, column=col_idx, value="WELL-001")
    wb.save(path)


def write_batch_example(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "BatchInput"
    _write_header_row(ws, BATCH_HEADERS)
    for row_idx, row in enumerate(example_rows(), start=2):
        for col_idx, header in enumerate(BATCH_HEADERS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
    wb.save(path)


def _autosize_columns(ws, max_width: int = 48) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)


def write_batch_results(
    path: Path,
    results: list[BatchRowResult],
    *,
    plots_dir: Path | None = None,
    plot_format: str = "png",
) -> Path | None:
    from openpyxl.styles import Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "BatchResults"
    headers = BATCH_HEADERS + RESULT_COLUMNS
    _write_header_row(ws, headers)
    bold = Font(bold=True)
    no_wrap = Alignment(wrap_text=False, vertical="center")
    # Columns after inputs: T_result, status, MC_*, error, intermediate_values
    col = {name: len(BATCH_HEADERS) + i + 1 for i, name in enumerate(RESULT_COLUMNS)}

    created_plots_dir: Path | None = None
    if plots_dir is not None:
        plots_dir.mkdir(parents=True, exist_ok=True)
        created_plots_dir = plots_dir

    for row_idx, result in enumerate(results, start=2):
        for col_idx, header in enumerate(BATCH_HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=result.inputs.get(header, ""))
            cell.alignment = no_wrap
        t_cell = ws.cell(row=row_idx, column=col["T_result"], value=result.t_result)
        t_cell.font = bold
        t_cell.alignment = no_wrap
        status_cell = ws.cell(row=row_idx, column=col["status"], value=result.status)
        status_cell.alignment = no_wrap
        if result.mc_stats:
            for key in MC_OUTPUT_COLUMNS:
                mc_cell = ws.cell(
                    row=row_idx, column=col[key], value=result.mc_stats.get(key)
                )
                mc_cell.alignment = no_wrap
        err_cell = ws.cell(
            row=row_idx,
            column=col["error"],
            value=str(result.error).replace("\r\n", " ").replace("\n", " ").replace("\r", " "),
        )
        err_cell.alignment = no_wrap
        inter_cell = ws.cell(
            row=row_idx,
            column=col["intermediate_values"],
            value=str(result.intermediate_values)
            .replace("\r\n", "; ")
            .replace("\n", "; ")
            .replace("\r", "; "),
        )
        inter_cell.alignment = no_wrap
        if (
            created_plots_dir is not None
            and result.mc_thicknesses
            and result.status == "OK"
        ):
            save_mc_pdf_cdf_plots(
                result.mc_thicknesses,
                created_plots_dir,
                result.well_id,
                result.t_number,
                plot_format=plot_format,
            )

    # Header row also no wrap.
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).alignment = no_wrap

    _autosize_columns(ws)
    wb.save(path)
    return created_plots_dir


def _safe_plot_stem(well_id: str, t_number: int) -> str:
    stem = re.sub(r"[^\w\-]+", "_", well_id.strip()) or "well"
    return f"{stem}_T{t_number}"


def save_mc_pdf_cdf_plots(
    thicknesses: list[float],
    plots_dir: Path,
    well_id: str,
    t_number: int,
    *,
    plot_format: str = "png",
) -> tuple[Path, Path]:
    import numpy as np
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    from matplotlib.figure import Figure

    if not thicknesses:
        raise ValueError("No Monte Carlo thickness samples to plot.")

    fmt = plot_format.lower().lstrip(".")
    if fmt not in {"png", "svg"}:
        raise ValueError(f"Unsupported Monte Carlo plot format: {plot_format!r}")

    stem = _safe_plot_stem(well_id, t_number)
    pdf_path = plots_dir / f"{stem}_pdf.{fmt}"
    cdf_path = plots_dir / f"{stem}_cdf.{fmt}"
    arr = np.asarray(thicknesses, dtype=float)
    n = len(arr)

    fig = Figure(figsize=(8, 4.5))
    FigureCanvasAgg(fig)
    ax = fig.add_subplot(111)
    pct_weight = 100.0 / n
    ax.hist(
        arr,
        bins=60,
        weights=[pct_weight] * n,
        density=False,
        color="#4C78A8",
        edgecolor="white",
        alpha=0.85,
        label="MC histogram",
    )
    try:
        from scipy.stats import gaussian_kde

        xs = np.linspace(arr.min(), arr.max(), 200)
        kde = gaussian_kde(arr)
        pdf = kde(xs)
        pdf = pdf / pdf.sum() * 100.0
        ax.plot(xs, pdf, color="#F58518", linewidth=2.0, label="KDE fit")
        ax.legend(loc="best")
    except Exception:
        pass
    ax.set_title(f"{well_id} — T{t_number} Monte Carlo PDF")
    ax.set_xlabel("Thickness")
    ax.set_ylabel("Percentage (%)")
    fig.tight_layout()
    fig.savefig(pdf_path, dpi=110, format=fmt)

    fig = Figure(figsize=(8, 4.5))
    FigureCanvasAgg(fig)
    ax = fig.add_subplot(111)
    sorted_vals = np.sort(arr)
    ys = np.arange(1, n + 1) / n * 100.0
    ax.plot(sorted_vals, ys, color="#4C78A8", linewidth=2.0)
    ax.set_title(f"{well_id} — T{t_number} Monte Carlo CDF")
    ax.set_xlabel("Thickness")
    ax.set_ylabel("Cumulative percentage (%)")
    ax.set_ylim(0, 100)
    fig.tight_layout()
    fig.savefig(cdf_path, dpi=110, format=fmt)
    return pdf_path, cdf_path


def batch_schema_json() -> str:
    from source.batch_schema import to_js_object

    return json.dumps(to_js_object(), indent=2, ensure_ascii=False)

from __future__ import annotations

import html
import io
import traceback
from contextlib import redirect_stderr, redirect_stdout
from pathlib import Path
from typing import Callable

from PySide6.QtCore import Qt
from PySide6.QtGui import QColor, QTextCharFormat, QTextCursor
from PySide6.QtWidgets import (
    QDoubleSpinBox,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QSplitter,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

# Shown on every σ (Monte Carlo uncertainty) spin box — short, plain language.
MONTE_CARLO_SIGMA_TOOLTIP = (
    "If your instrument quotes ±ε for this reading (e.g. ±2°), it is\n"
    "common to use σ = ε ÷ 2. While σ = 0 yields a single thickness output,\n"
    "non-zero σ explores a range of possible values accounting for the\n"
    "inaccuracies resulting from instrument errors.\n\n"
    "For more accurate calculations, use the confidence level stated in\n"
    "your manual (most field instruments quote ±ε for a 95% interval).\n"
    "Then set σ = ε ÷ z, where z is the z-score for that level:\n"
    "z ≈ 1.645 (90%), 1.960 (95%), 2.576 (99%). The rule σ = ε ÷ 2\n"
    "above is the 95% case rounded, since z for 95% is close to 2."
)


class _DoubleSpinBoxNoTrailingZeros(QDoubleSpinBox):
    """
    Same precision and stepping as QDoubleSpinBox, but the line edit does not
    pad with trailing zeros (e.g. 20 instead of 20.0000).
    """

    def textFromValue(self, value: float) -> str:
        if value != value:  # NaN
            return super().textFromValue(value)
        if value == 0.0:
            return "0"
        d = self.decimals()
        s = f"{value:.{d}f}".rstrip("0").rstrip(".")
        if s in ("", "-", "-0"):
            return "0"
        return s


class ModelTab(QWidget):
    """
    Generic tab layout:
    - input and output side by side
    - calculate / clear row
    - resizable splitters: input | output; stdout/stderr below
    """

    def __init__(
        self,
        title: str,
        on_calculate: Callable[["ModelTab"], None],
        export_basename: str,
    ) -> None:
        super().__init__()
        self.on_calculate = on_calculate
        self.export_basename = export_basename
        self.inputs: dict[str, QDoubleSpinBox] = {}
        self.std_inputs: dict[str, QDoubleSpinBox] = {}
        self._mc_thicknesses: list[float] | None = None
        self._mc_title: str = ""
        self._xlsx_input_columns: list[tuple[str, float, float]] | None = None
        self._xlsx_output_rows: list[tuple[str, float | int]] | None = None
        self._xlsx_mc_rows: list[tuple[str, float | int]] | None = None
        self.mc_save_fn: Callable[[list[float], str, str, str], None] | None = None
        self._log_normal_fg = QColor(0, 0, 0)
        self._build_ui(title)

    def _build_ui(self, title: str) -> None:
        root_layout = QVBoxLayout()
        self.setLayout(root_layout)

        self._title_label = QLabel(title)

        self.input_group = QGroupBox("Input")
        self.input_form = QFormLayout()
        self.input_group.setLayout(self.input_form)

        self.output_group = QGroupBox("Output")
        output_layout = QVBoxLayout()
        self.output_text = QTextEdit()
        self.output_text.setReadOnly(True)
        output_layout.addWidget(self.output_text)

        mc_row = QHBoxLayout()
        self._save_excel_btn = QPushButton("Save Results to Excel")
        self._save_excel_btn.clicked.connect(self._on_save_excel)
        self._save_excel_btn.setVisible(False)
        self._mc_save_png_btn = QPushButton("Save MC plot (PNG)")
        self._mc_save_svg_btn = QPushButton("Save MC plot (SVG)")
        self._mc_save_png_btn.setVisible(False)
        self._mc_save_svg_btn.setVisible(False)
        self._mc_save_png_btn.clicked.connect(self._on_save_mc_png)
        self._mc_save_svg_btn.clicked.connect(self._on_save_mc_svg)
        mc_row.addWidget(self._save_excel_btn)
        mc_row.addWidget(self._mc_save_png_btn)
        mc_row.addWidget(self._mc_save_svg_btn)
        mc_row.addStretch()
        output_layout.addLayout(mc_row)

        self.output_group.setLayout(output_layout)

        content_splitter = QSplitter(Qt.Orientation.Horizontal)
        content_splitter.setChildrenCollapsible(False)
        content_splitter.addWidget(self.input_group)
        content_splitter.addWidget(self.output_group)
        content_splitter.setStretchFactor(0, 1)
        content_splitter.setStretchFactor(1, 1)

        button_row = QHBoxLayout()
        self.calculate_button = QPushButton("Calculate")
        self.calculate_button.clicked.connect(self._run_calculation)
        self.clear_button = QPushButton("Clear")
        self.clear_button.clicked.connect(self._clear_all)
        button_row.addWidget(self.calculate_button)
        button_row.addWidget(self.clear_button)
        button_row.addStretch()

        top_area = QWidget()
        top_layout = QVBoxLayout(top_area)
        top_layout.addWidget(self._title_label)
        top_layout.addWidget(content_splitter, stretch=1)
        top_layout.addLayout(button_row)

        logs_group = QGroupBox("Stdout / Stderr")
        logs_layout = QVBoxLayout()
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setPlaceholderText(
            "Combined logs. stderr entries are shown in red."
        )
        logs_layout.addWidget(self.log_text)
        logs_group.setLayout(logs_layout)

        main_splitter = QSplitter(Qt.Orientation.Vertical)
        main_splitter.setChildrenCollapsible(False)
        main_splitter.addWidget(top_area)
        main_splitter.addWidget(logs_group)
        main_splitter.setStretchFactor(0, 1)
        main_splitter.setStretchFactor(1, 1)

        root_layout.addWidget(main_splitter)

        content_splitter.setSizes([520, 520])
        main_splitter.setSizes([520, 280])

    def add_float_input(
        self,
        key: str,
        label: str,
        decimals: int = 4,
        minimum: float = -100000.0,
        maximum: float = 100000.0,
        step: float = 0.1,
        default: float = 0.0,
    ) -> None:
        value_box = _DoubleSpinBoxNoTrailingZeros()
        value_box.setDecimals(decimals)
        value_box.setRange(minimum, maximum)
        value_box.setSingleStep(step)
        value_box.setValue(default)

        std_box = _DoubleSpinBoxNoTrailingZeros()
        std_box.setDecimals(decimals)
        std_box.setRange(0.0, maximum - minimum)
        std_box.setSingleStep(step)
        std_box.setValue(0.0)
        std_box.setPrefix("σ=")
        std_box.setToolTip(MONTE_CARLO_SIGMA_TOOLTIP)

        row_widget = QWidget()
        row_layout = QHBoxLayout(row_widget)
        row_layout.setContentsMargins(0, 0, 0, 0)
        row_layout.addWidget(value_box, stretch=3)
        row_layout.addWidget(std_box, stretch=2)

        self.inputs[key] = value_box
        self.std_inputs[key] = std_box
        self.input_form.addRow(label, row_widget)

    def apply_theme(self, dark: bool) -> None:
        self._log_normal_fg = (
            QColor(0xE8, 0xE8, 0xE8) if dark else QColor(0, 0, 0)
        )
        title_color = "#e8e8e8" if dark else "#000000"
        self._title_label.setStyleSheet(
            f"font-size: 16px; font-weight: 600; color: {title_color}; "
            "background: transparent;"
        )
        # Geometry warning call-outs use class geometry-warnings; base rules would
        # otherwise paint li/span/b inside them with the normal output color.
        _warn_colors = (
            "div.geometry-warnings, div.geometry-warnings span, "
            "div.geometry-warnings b, div.geometry-warnings li { color: #b91c1c; } "
        )
        if dark:
            self.output_text.document().setDefaultStyleSheet(
                "body, p, div, span, b, li { color: #e8e8e8; } "
                "a { color: #8cb4ff; } "
                + _warn_colors
            )
        else:
            self.output_text.document().setDefaultStyleSheet(
                "body, p, div, span, b, li { color: #000000; } "
                "a { color: #0066cc; } "
                + _warn_colors
            )

    def value(self, key: str) -> float:
        return self.inputs[key].value()

    def std(self, key: str) -> float:
        return self.std_inputs[key].value()

    def set_output(self, text: str, is_html: bool = False) -> None:
        # Clear first so embedded resources (e.g. Monte Carlo PNGs) do not persist
        # across runs when the next output omits them.
        self.output_text.clear()
        self.set_monte_carlo_export_state(None, "")
        self._xlsx_input_columns = None
        self._xlsx_output_rows = None
        self._xlsx_mc_rows = None
        self._save_excel_btn.setVisible(False)
        if is_html:
            self.output_text.setHtml(text)
        else:
            self.output_text.setPlainText(text)

    def set_monte_carlo_export_state(
        self,
        thicknesses: list[float] | None,
        title: str = "",
    ) -> None:
        self._mc_thicknesses = list(thicknesses) if thicknesses else None
        self._mc_title = title
        visible = bool(self._mc_thicknesses)
        self._mc_save_png_btn.setVisible(visible)
        self._mc_save_svg_btn.setVisible(visible)

    def set_export_snapshot_sections(
        self,
        input_columns: list[tuple[str, float, float]],
        output_rows: list[tuple[str, float | int]],
        mc_rows: list[tuple[str, float | int]] | None,
    ) -> None:
        self._xlsx_input_columns = list(input_columns)
        self._xlsx_output_rows = list(output_rows)
        self._xlsx_mc_rows = list(mc_rows) if mc_rows is not None else None
        self._save_excel_btn.setVisible(self._xlsx_export_ready())

    def _xlsx_export_ready(self) -> bool:
        return (
            self._xlsx_input_columns is not None
            and len(self._xlsx_input_columns) > 0
            and self._xlsx_output_rows is not None
            and len(self._xlsx_output_rows) > 0
        )

    def _default_export_path(self, extension: str) -> str:
        ext = extension.lstrip(".")
        return str(Path.cwd().resolve() / f"{self.export_basename}.{ext}")

    def _on_save_excel(self) -> None:
        if not self._xlsx_export_ready():
            QMessageBox.information(
                self,
                "Save Results to Excel",
                "Run Calculate first to save results.",
            )
            return
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Results to Excel",
            self._default_export_path("xlsx"),
            "Excel Workbook (*.xlsx)",
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            self._write_results_xlsx(
                path,
                self.export_basename,
                self._xlsx_input_columns,
                self._xlsx_output_rows,
                self._xlsx_mc_rows,
            )
            print(f"Saved Excel workbook to {path}")
        except OSError as exc:
            QMessageBox.warning(self, "Save Results to Excel", str(exc))
        except Exception as exc:
            QMessageBox.warning(
                self,
                "Save Results to Excel",
                f"Could not write workbook: {exc}",
            )

    @staticmethod
    def _write_results_xlsx(
        path: str,
        model_basename: str,
        input_columns: list[tuple[str, float, float]] | None,
        output_rows: list[tuple[str, float | int]] | None,
        mc_rows: list[tuple[str, float | int]] | None,
    ) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        if not input_columns or not output_rows:
            raise ValueError("Missing export data")

        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        bold = Font(bold=True)
        r = 1

        ws.cell(row=r, column=1, value="Model").font = bold
        ws.cell(row=r, column=2, value=model_basename).font = bold
        r += 2  # row 2 blank; Inputs start on row 3

        ws.cell(row=r, column=1, value="Inputs").font = bold
        r += 1
        ws.cell(row=r, column=1, value="Field").font = bold
        ws.cell(row=r, column=2, value="Value").font = bold
        ws.cell(row=r, column=3, value="Sigma").font = bold
        r += 1
        for name, val, sigma in input_columns:
            ws.cell(row=r, column=1, value=name)
            ws.cell(row=r, column=2, value=val)
            ws.cell(row=r, column=3, value=sigma)
            r += 1
        r += 1

        ws.cell(row=r, column=1, value="Outputs").font = bold
        r += 1
        ws.cell(row=r, column=1, value="Field").font = bold
        ws.cell(row=r, column=2, value="Value").font = bold
        r += 1
        for name, val in output_rows:
            ws.cell(row=r, column=1, value=name)
            ws.cell(row=r, column=2, value=val)
            r += 1

        if mc_rows:
            r += 1
            ws.cell(row=r, column=1, value="Monte Carlo").font = bold
            r += 1
            ws.cell(row=r, column=1, value="Field").font = bold
            ws.cell(row=r, column=2, value="Value").font = bold
            r += 1
            for name, val in mc_rows:
                ws.cell(row=r, column=1, value=name)
                ws.cell(row=r, column=2, value=val)
                r += 1

        wb.save(path)

    def _on_save_mc_png(self) -> None:
        self._save_mc_plot_dialog("png", "PNG Images (*.png)")

    def _on_save_mc_svg(self) -> None:
        self._save_mc_plot_dialog("svg", "SVG Vector Graphics (*.svg)")

    def _save_mc_plot_dialog(self, fmt: str, name_filter: str) -> None:
        if not self._mc_thicknesses or not self.mc_save_fn:
            return
        path, _ = QFileDialog.getSaveFileName(
            self,
            f"Save Monte Carlo histogram ({fmt.upper()})",
            self._default_export_path(fmt),
            name_filter,
        )
        if not path:
            return
        ext = f".{fmt}"
        if not path.lower().endswith(ext):
            path += ext
        self.mc_save_fn(self._mc_thicknesses, self._mc_title, fmt, path)

    def _clear_all(self) -> None:
        self.output_text.clear()
        self.set_monte_carlo_export_state(None, "")
        self._xlsx_input_columns = None
        self._xlsx_output_rows = None
        self._xlsx_mc_rows = None
        self._save_excel_btn.setVisible(False)
        self.log_text.clear()

    def _append_log(self, text: str, is_error: bool) -> None:
        if not text:
            return
        cursor = self.log_text.textCursor()
        cursor.movePosition(QTextCursor.End)
        text_format = QTextCharFormat()
        text_format.setForeground(
            QColor("red") if is_error else self._log_normal_fg
        )
        cursor.insertText(text, text_format)
        if not text.endswith("\n"):
            cursor.insertText("\n", text_format)
        self.log_text.setTextCursor(cursor)
        self.log_text.ensureCursorVisible()

    def _set_combined_logs(self, stdout_text: str, stderr_text: str) -> None:
        self.log_text.clear()
        stdout_lines = stdout_text.splitlines(keepends=True)
        stderr_lines = stderr_text.splitlines(keepends=True)
        max_len = max(len(stdout_lines), len(stderr_lines))
        for index in range(max_len):
            if index < len(stdout_lines):
                self._append_log(stdout_lines[index], is_error=False)
            if index < len(stderr_lines):
                self._append_log(stderr_lines[index], is_error=True)

    def _run_calculation(self) -> None:
        stdout_buffer = io.StringIO()
        stderr_buffer = io.StringIO()
        self.log_text.clear()
        try:
            with redirect_stdout(stdout_buffer), redirect_stderr(stderr_buffer):
                self.on_calculate(self)
        except Exception as exc:  # pragma: no cover
            traceback.print_exc(file=stderr_buffer)
            msg = html.escape(f"Calculation failed: {exc}")
            self.set_output(
                f'<p style="color:#b91c1c;">{msg}</p>',
                is_html=True,
            )
        finally:
            self._set_combined_logs(
                stdout_buffer.getvalue(),
                stderr_buffer.getvalue(),
            )
